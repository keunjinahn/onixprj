print ("module [backend.api_common] loaded")

import hashlib
from flask import make_response, jsonify, request, json, send_from_directory, g
from flask_restless import ProcessingException
from flask_restful import reqparse
from datetime import datetime, timedelta
import os
import json
from functools import wraps
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend import app, login_manager
from backend_model.table_fireprj import *
from backend import manager
db = DBManager.db
import random

manager.create_api(FireSensorTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

def post_repeter_log(result=None, **kw):
    res = result['objects']
    for repeter in res:
        receiver = FireReceiverTbl.query.filter_by(fk_customer_idx=repeter['fk_customer_idx']).filter_by(receiver_id=repeter['receiver_id']).first()
        if receiver is not None :
            repeter['receiver']['receiver_type'] = receiver.receiver_type

manager.create_api(FireRepeaterTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='repeater'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , postprocessors={
                        'GET_MANY': [post_repeter_log]
                   })

manager.create_api(FireReceiverTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='receiver'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

def post_sensor_log(result=None, **kw):
    res = result['objects']
    for sensor in res:
        receiver = FireReceiverTbl.query.filter_by(fk_customer_idx=sensor['fk_customer_idx']).filter_by(receiver_id=sensor['receiver_id']).first()
        if receiver is not None :
            sensor['receiver']['receiver_type'] = receiver.receiver_type

manager.create_api(FireSensorTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor_event'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , postprocessors={
                        'GET_MANY': [post_sensor_log]
                   })

manager.create_api(EventTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='event_list'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

manager.create_api(EventTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sensor_analysis'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

def prePasswdUpdate(input_params=None, **kw):
    if 'user_pwd' in kw['data']:
        kw['data']['user_pwd'] = password_encoder_512(kw['data']['user_pwd'])
        
manager.create_api(UserTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='user'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [prePasswdUpdate],
                        'PATCH_SINGLE': [prePasswdUpdate],
                   })            

manager.create_api(CustomerTbl
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='customer'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True)

# manager.create_api(EventLogTbl
#                    , results_per_page=10000
#                    , url_prefix='/api/v1'
#                    , collection_name='event_log_list'
#                    , methods=['GET', 'DELETE', 'PATCH', 'POST']
#                    , allow_patch_many=True)

@app.route('/make-data/event-list', methods=['GET'])
def write_event_list_api():
    file_path = 'C:\\Users\\ansieun\\project\\fireprj\\doc\\화재감시시스템 이벤트 리스트_20240326.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Sheet1')
    
    data = []
    for i, row in enumerate(sheet.rows):
        row_data = []
        for j, col in enumerate(row):
            if j < 1: continue
            row_data.append(col.value)
        if row_data[0] is not None: data.append(row_data)
    print(len(data))
    for i, row in enumerate(data):
        if not isinstance(row[0], int):
            del data[i]
            continue
    print(len(data))

    table_list = ['event_idx', 'system_id_c', 'repeater_id_c', 'sensor_id_c', 'sensor_value_c', 'inout_id_c', 'event_desc', 'event_category']
    for i, row in enumerate(data):
        obj = EventTbl()
        for j, col in enumerate(row):
            setattr(obj, table_list[j], col)
        db.session.add(obj)
    db.session.commit()
    return make_response(jsonify(''), 200)

@app.route('/make-data/generating', methods=['POST'])
def generating_api():
    generate_option = json.loads(request.data)

    if 'receiver' in generate_option:
        receiver_len = generate_option['receiver']
        for i in range(receiver_len):
            receiver_id = i+1
            db.session.add(FireReceiverTbl(
                receiver_idx    = f'00001_001_{receiver_id:0>3}',
                fk_customer_idx = 1,                        #고객식별자: 1
                receiver_type   = 1,                        #수신기 타입: 1
                receiver_id     = receiver_id
            ))
        db.session.commit()

    if 'repeater' in generate_option:
        repeater_len = generate_option['repeater']
        receiver_data = FireReceiverTbl.query.all()
        for i, receiver in enumerate(receiver_data):
            for j in range(repeater_len):
                repeater_id = (i*repeater_len) + (j+1)
                db.session.add(FireRepeaterTbl(
                    repeater_idx    = f'00001_{receiver.receiver_id:0>3}_000_{repeater_id:0>3}',
                    fk_customer_idx = 1,                    #고객식별자: 1
                    receiver_id     = receiver.receiver_id,
                    system_id       = 0,                    #계통번호: 0
                    repeater_id     = repeater_id,
                ))
        db.session.commit()

    if 'sensor' in generate_option:
        sensor_len = generate_option['sensor']
        repeater_data = FireRepeaterTbl.query.all()
        for i, repeater in enumerate(repeater_data):
            for j in range(sensor_len):
                sensor_id = (i*sensor_len) + (j+1)
                db.session.add(FireSensorTbl(
                    sensor_idx      = repeater.repeater_idx + f'_{sensor_id:0>3}',
                    fk_customer_idx = 1,                    #고객식별자: 1
                    receiver_id     = repeater.receiver_id,
                    system_id       = 0,                    #계통번호: 0
                    repeater_id     = repeater.repeater_id,
                    sensor_id       = sensor_id,
                ))
        db.session.commit()

    return make_response(jsonify(''), 200)

@app.route('/api/v1/login', methods=['POST'])
def login_api():
    data = json.loads(request.data)
    result = ''
    print("111")
    if data['username'] is not None and data['password'] is not None:
        loginuser = db.session.query(UserTbl).filter(UserTbl.user_id == data["username"]).first()

        if loginuser is None:
            result = {'status': False, 'reason': 1}  # ID 없음
        else:
            if loginuser.user_pwd != password_encoder_512(data["password"]):
                result = {'status': False, 'reason': 2} # PW 틀림
                print("222")
            else:  # Login 성공
                if loginuser.user_status == 2:
                    result = {'status': False, 'reason': 3}  # Activation 안됨
                else:
                    print("333")
                    loginuser.token = generate_token(data['username'])
                    db.session.query(UserTbl).filter(UserTbl.user_id == data["username"])\
                        .update(dict(token=loginuser.token))
                    db.session.commit()

                    result = {'status': True, 'reason': 0, 'user': loginuser.serialize()}
    print("444:",result)
    return make_response(jsonify(result), 200)

@app.route("/api/v1/logout", methods=["POST"])
def logout_api():
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]
    result = ''
    if token is None:
        print("token is none")

    loginUser = UserTbl.query.filter_by(token=token).first()
    if loginUser is None:
        print("user is none")

    return make_response(jsonify(result), 200)

def generate_token(userID):
    m = hashlib.sha1()

    m.update(userID.encode('utf-8'))
    m.update(datetime.now().isoformat().encode('utf-8'))

    return m.hexdigest()

def check_token(search_params=None, **kw):
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]
    if token is None:
        raise ProcessingException(description="Not Authorized", code=410)
    user = UserTbl.query.filter_by(token=token).first()
    if user is None:
        raise ProcessingException(description="Not Authorized", code=411)

def check_token_single(search_params=None, **kw):
    parser = reqparse.RequestParser()
    parser.add_argument("token", type=str, location="headers")
    token = parser.parse_args()["token"]

    if token is None:
        raise ProcessingException(description="Not Authorized", code=410)

    user = UserTbl.query.filter_by(token=token).first()
    if user is None:
        raise ProcessingException(description="Not Authorized", code=411)

def password_encoder_512(password):
    h = hashlib.sha512()
    h.update(password.encode('utf-8'))
    return h.hexdigest()

def parse_dbdata_for_excel_format(page, values):
    if page == 'sensor_event':
        tmp = ['정상' if v == 1 else '이상' for v in values[6:10]]
        tmp.append(values[10].strftime("%Y-%m-%d %H:%M:%S") if values[10] is not None else None)
        values = values[:6]
    elif page == 'repeater_event':
        tmp = ['정상' if v == 1 else '이상' for v in values[5:9]]
        tmp.append(values[9].strftime("%Y-%m-%d %H:%M:%S") if values[9] is not None else None)
        values = values[:5]

    if page == 'sensor_manage' or page == 'repeater_manage':
        values = values[:-5]
    if 'manage' in page or '_event' in page:
        values[1] = str(values[1]).zfill(5)
        values[2:] = [str(v).zfill(3) for v in values[2:]]
    if '_event' in page:
        values += tmp

    if page == 'event_list':
        values[1:6] = [str(v).zfill(3) for v in values[1:6]]
    if page == 'user':
        del values[-1]
        values[1] = '****'
        values[3] = '사용' if values[3] == 1 else '미사용'
        values[4] = '관리' if values[4] == 1 else '조회'
    return values

@app.route('/api/v1/make_excel', methods=['POST'])
def make_excel_api():
    data = json.loads(request.data)
    page = data['page_name']
    headers = data['headers']
    if '_event' in page:
        headers = headers[1:] + [headers[0]]
    db_router = {
        'sensor_manage': FireSensorTbl(),
        'repeater_manage': FireRepeaterTbl(),
        'receiver_manage': FireReceiverTbl(),
        'sensor_event': FireSensorTbl(),
        'repeater_event': FireRepeaterTbl(),
        'event_list': EventTbl(),
        'user': UserTbl(),
        'crm': CustomerTbl(),
    }
    target_db = db_router[page]
    db_data = target_db.query.all()

    target_path = './download/'
    file_name = f'{page} {datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    wb = Workbook()
    ws = wb.active

    ws.append(headers) #헤더 삽입
    for row in db_data: #내용 삽입
        values = [getattr(row, col.name) for col in row.__table__.columns][1:]
        values = parse_dbdata_for_excel_format(page, values)
        ws.append(values)

    bd_thin = Side(border_style='thin', color='000000')
    bd_all = Border(bd_thin, bd_thin, bd_thin, bd_thin)
    for row in ws.rows: #테두리, 정렬
        for cell in row:
            cell.border = bd_all
            cell.alignment = Alignment(horizontal='left')
    for cell in ws['1']: #헤더 스타일 지정
        cell.fill = PatternFill(fill_type='solid', fgColor=Color('a0a0a0'))
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(cell.col_idx)].width = 15 #열 크기 지정
    if page == 'sensor_manage': ws.column_dimensions['A'].width = 23
    if page == 'repeater_manage': ws.column_dimensions['A'].width = 17
    if page == 'receiver_manage': ws.column_dimensions['A'].width = 14
    if page == 'event_list': ws.column_dimensions['G'].width = 50
    if page == 'crm': ws.column_dimensions['C'].width = 50
    if page == 'sensor_event':
        ws.column_dimensions['A'].width = 23
        ws.column_dimensions['K'].width = 18
    if page == 'repeater_event':
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['J'].width = 18

    wb.save(target_path+file_name)
    return make_response(jsonify({"filename":file_name}), 200)

@app.route('/api/v1/save_excel/<filename>', methods=['GET'])
def save_excel_api(filename):
    upload_path = os.getcwd() + '/download'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/sensor-log-chart', methods=['GET'])
def sensor_log_chart_api():
    sensor_id = request.args.get('sensor_id')
    data = EventLogTbl.query.filter_by(sensor_id=sensor_id)\
        .order_by(EventLogTbl.event_datetime.desc())\
        .offset(0).limit(10).all()
    values = []
    dates = []
    for d in data:
        values.append(float(d.sensor_value))
        dates.append(d.event_datetime)
    values = values[::-1]
    dates = dates[::-1]

    obj = {
        "objects": values,
        "dates": dates
    }
    return make_response(jsonify(obj), 200)

@app.route('/api/v1/insert_test_sensor_data', methods=['POST'])
def insert_test_sensor_data_api():
    sensor_list = FireSensorTbl.query.all()
    receiver_list = FireReceiverTbl.query.all()
    event_count = EventTbl.query.count()
    str_start_date = "2024-01-01 00:00:00"
    start_date =  datetime.strptime(str_start_date,'%Y-%m-%d %H:%M:%S')
    for idx in range(24 * 3600): #24 * 3600
        c_date = start_date + timedelta(seconds=idx)
        for sensor in sensor_list :
            new_data = EventLogTbl()
            new_data.fk_customer_idx = sensor.fk_customer_idx
            new_data.event_id = random.randint(1,event_count)
            receiver = list(filter(lambda x:x.receiver_id == sensor.receiver_id,receiver_list))[0]
            new_data.receiver_type = receiver.receiver_type if receiver is not None else 0
            new_data.receiver_id = sensor.receiver_id
            new_data.system_id = sensor.system_id
            new_data.repeater_id = sensor.repeater_id
            new_data.sensor_id = sensor.sensor_id
            new_data.sensor_value = round(random.uniform(0, 10),4)
            new_data.inout_id = random.randint(1,2)
            new_data.event_datetime = c_date
            db.session.add(new_data)
        db.session.commit()
        print("idx :",idx, "date : ",c_date)
        
    return make_response(jsonify({}), 200)

@app.route('/api/v1/sensor_data_range', methods=['POST'])
def sensor_data_range_api():
    print("sensor_data_range start...")
    input = json.loads(request.data)
    range_value = int(input['range_value'])
    sensor_id = input['sensor_id']
    str_start_date = "2024-01-01 " + datetime.now().strftime('%H:%M:%S')
    start_date =  datetime.strptime(str_start_date,'%Y-%m-%d %H:%M:%S')
    def get_start_date(start_date,range_value):
        s_date = start_date - timedelta(hours=1)            
        if range_value == 3 :
            s_date = start_date - timedelta(minutes=1)
        elif range_value == 2 :
            s_date = start_date - timedelta(minutes=10)
        elif range_value == 1 :
            s_date = start_date - timedelta(hours=1)            
        return s_date
    print("input: ",input)
    start_date = get_start_date(start_date,range_value)
    str_end_date = "2024-01-01 " + datetime.now().strftime('%H:%M:%S')
    str_start_date = start_date.strftime("%Y-%m-%d %H:%M:%S")
    
    print("start : ",str_start_date, ", end : ",str_end_date)
    sensor_rt_data_list = EventLogTbl.query \
        .filter(EventLogTbl.sensor_id == sensor_id) \
        .filter(EventLogTbl.event_datetime >= str_start_date) \
        .filter(EventLogTbl.event_datetime <= str_end_date) \
        .all()
    print("sensor_rt_data_list len :", len(sensor_rt_data_list))
    sensor_value_list = []    
    for sensor_rt_data in sensor_rt_data_list:
        event_date = datetime.now().strftime('%Y-%m-%d') + " " +  sensor_rt_data.event_datetime.strftime('%H:%M:%S')
        sensor_value_list.append({"data":sensor_rt_data.sensor_value,"event_datetime":event_date})
    result_obj = {
        "sensor_rt_data_list":sensor_value_list,
    }
    return make_response(jsonify(result_obj), 200)    

@app.route('/api/v1/check_passwd', methods=['POST'])
def check_passwd_api():
    print("check_passwd...")
    input = json.loads(request.data)
    user_pwd = input['user_pwd']
    user_id = input['user_id']
    find_user = UserTbl.query.filter_by(user_id=user_id).filter_by(user_pwd=password_encoder_512(user_pwd)).first()
    if find_user is not None :
        return make_response(jsonify({"result":1}), 200)   
    return make_response(jsonify({"result":0}), 200)   


# 엑셀 파일 받아오기
@app.route('/api/v1/upload', methods=['POST'])
def except_event_xl():
    if 'files' not in request.files:
        return jsonify({"error":"No files part"}, 400)
    
    files = request.files.getlist('files')
    first = files[0]
    df = pd.read_excel(first, engine='openpyxl')

    xl_cols = ['GPS수집시간', '위도', '경도', '충격', '속도', '공회전', '배터리', '온도', '습도']
    table_map = ['eventStartTime', 'eventEndTime', 'latitude', 'longitude', 'impact', 'speed', 'idle',  'battery', 'temperature', 'humidity', 'createdTime']
    
    # 임시
    df.insert(1, 'new_time', df['GPS수집시간'])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df['now'] = now

    df.columns = table_map
    # print(df.head())

    
    df.to_sql('tbeventlog', con=db.engine, if_exists='append', index=False)

    return make_response(jsonify({"result":0}), 200)